#!/usr/bin/env python3
"""
generate_report.py — creates a two-sheet Excel report matching the
uploaded Board_Test_Report format exactly.
One sheet per board: "Board1 - PHYTEC AM335x" and "Board2 - SAMA5D2"

Usage: python3 generate_report.py <build_number> <build_url> <status>
"""

import sys, os, glob, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Colours matching the uploaded template
# ---------------------------------------------------------------------------
DARK_BLUE   = "1F4E79"   # title bar, test header row, output row
MED_BLUE    = "2E75B6"   # TEST RESULTS section header
LIGHT_GREY  = "F2F2F2"   # alternating metadata rows
PASS_GREEN  = "C6EFCE"   # PASSED cell background
PASS_FONT   = "1E7B34"   # PASSED cell font colour
FAIL_RED    = "FFC7CE"   # FAILED cell background
FAIL_FONT   = "9C0006"   # FAILED cell font colour
WHITE       = "FFFFFF"

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, color=WHITE, size=11):
    return Font(bold=bold, color=color, size=size)

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def read_log(path_pattern):
    """Read first matching log file, return content or empty string."""
    files = glob.glob(path_pattern)
    if not files:
        return ""
    try:
        return open(files[0]).read().strip()
    except Exception:
        return ""

def result_of(log_content):
    """Determine PASS/FAIL from log content."""
    if not log_content:
        return "SKIPPED"
    if "❌" in log_content or "timed out" in log_content or "FAIL" in log_content:
        return "FAILED ❌"
    return "PASSED ✅"

def build_sheet(wb, sheet_name, board_id, board_label, board_port,
                build_number, build_time, status):
    ws = wb.create_sheet(title=sheet_name)

    # ------------------------------------------------------------------
    # Row 1 — Title
    # ------------------------------------------------------------------
    ws.merge_cells("A1:E1")
    ws["A1"] = "JENKINS BOARD TEST REPORT"
    ws["A1"].fill      = fill(DARK_BLUE)
    ws["A1"].font      = font(bold=True, color=WHITE, size=16)
    ws["A1"].alignment = align("center")
    ws.row_dimensions[1].height = 27.75

    # ------------------------------------------------------------------
    # Rows 2-6 — Metadata
    # ------------------------------------------------------------------
    meta = [
        ("Build Number", f"#{build_number}"),
        ("Build Time",   build_time),
        ("Board Port",   board_port),
        ("Board User",   "root"),
        ("Agent Used",   "slave-1"),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        bg = LIGHT_GREY if i % 2 == 0 else WHITE
        ws.merge_cells(f"B{i}:E{i}")
        ws[f"A{i}"] = label
        ws[f"A{i}"].fill      = fill(bg)
        ws[f"A{i}"].font      = Font(bold=True, size=11, color="000000")
        ws[f"A{i}"].alignment = align("left", wrap=False)
        ws[f"B{i}"] = value
        ws[f"B{i}"].fill      = fill(bg)
        ws[f"B{i}"].font      = Font(bold=False, size=11, color="000000")
        ws[f"B{i}"].alignment = align("left", wrap=False)
        ws.row_dimensions[i].height = 15

    # Row 7 blank
    ws.row_dimensions[7].height = 8

    # ------------------------------------------------------------------
    # Row 8 — TEST RESULTS section header
    # ------------------------------------------------------------------
    ws.merge_cells("A8:E8")
    ws["A8"] = "TEST RESULTS"
    ws["A8"].fill      = fill(MED_BLUE)
    ws["A8"].font      = font(bold=True, color=WHITE, size=13)
    ws["A8"].alignment = align("center")
    ws.row_dimensions[8].height = 21.75

    # ------------------------------------------------------------------
    # Row 9 — Stage headers
    # ------------------------------------------------------------------
    stages = [
        "1. Board Reachable",
        "2. Serial Session",
        "3. GPIO Discovery",
        "4. Peripheral Test",
        "5. Functional Test",
    ]
    cols = "ABCDE"
    for col, stage in zip(cols, stages):
        c = ws[f"{col}9"]
        c.value     = stage
        c.fill      = fill(DARK_BLUE)
        c.font      = font(bold=True, color=WHITE, size=11)
        c.alignment = align("center")
    ws.row_dimensions[9].height = 30

    # ------------------------------------------------------------------
    # Read log files
    # ------------------------------------------------------------------
    logs = {
        "A": read_log(f"test-logs/{board_id}/A_*.log"),
        "B": read_log(f"test-logs/{board_id}/B_*.log"),
        "C": read_log(f"test-logs/{board_id}/C_*.log"),
        "D": read_log(f"test-logs/{board_id}/D_*.log"),
        "E": read_log(f"test-logs/{board_id}/E_*.log"),
    }

    # ------------------------------------------------------------------
    # Row 10 — Log output content
    # ------------------------------------------------------------------
    for col, log in logs.items():
        c = ws[f"{col}10"]
        clean_log = ''.join(ch for ch in (log if log else "(no log)") if ord(ch) >= 32 or ch in '\t\n\r')
        c.value     = clean_log
        c.fill      = fill(WHITE)
        c.font      = Font(bold=False, color="000000", size=10)
        c.alignment = align("left", wrap=True)
    ws.row_dimensions[10].height = 345.5

    # ------------------------------------------------------------------
    # Row 11 — PASSED / FAILED per stage
    # ------------------------------------------------------------------
    results = {}
    for col, log in logs.items():
        r = result_of(log)
        results[col] = r
        bg  = PASS_GREEN if "PASSED" in r else (FAIL_RED if "FAILED" in r else "D9D9D9")
        fc  = PASS_FONT  if "PASSED" in r else (FAIL_FONT if "FAILED" in r else "666666")
        c = ws[f"{col}11"]
        c.value     = r
        c.fill      = fill(bg)
        c.font      = Font(bold=True, color=fc, size=11)
        c.alignment = align("center")
    ws.row_dimensions[11].height = 13.8

    # Row 12 blank
    ws.row_dimensions[12].height = 13.8

    # ------------------------------------------------------------------
    # Row 13 — Overall status
    # ------------------------------------------------------------------
    ws.merge_cells("A13:E13")
    all_pass = all("PASSED" in v for v in results.values())
    overall  = "STATUS: ALL TESTS PASSED ✅" if all_pass else "STATUS: SOME TESTS FAILED ❌"
    bg  = PASS_GREEN if all_pass else FAIL_RED
    fc  = PASS_FONT  if all_pass else FAIL_FONT
    ws["A13"] = overall
    ws["A13"].fill      = fill(bg)
    ws["A13"].font      = Font(bold=True, color=fc, size=13)
    ws["A13"].alignment = align("center")
    ws.row_dimensions[13].height = 25.5

    # ------------------------------------------------------------------
    # Column widths — match uploaded template
    # ------------------------------------------------------------------
    ws.column_dimensions["A"].width = 58.32
    ws.column_dimensions["B"].width = 51.04
    ws.column_dimensions["C"].width = 50.71
    ws.column_dimensions["D"].width = 25.91
    ws.column_dimensions["E"].width = 60.42

    return ws


def main():
    build_number = sys.argv[1] if len(sys.argv) > 1 else "0"
    build_url    = sys.argv[2] if len(sys.argv) > 2 else ""
    status       = sys.argv[3] if len(sys.argv) > 3 else "UNKNOWN"

    now       = datetime.datetime.now()
    build_time = now.strftime("%A %d %B %Y %I:%M:%S %p IST")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    build_sheet(wb,
        sheet_name   = "Board1 - PHYTEC AM335x",
        board_id     = "board1",
        board_label  = "PHYTEC AM335x",
        board_port   = "/dev/ttyUSB0 @ 115200",
        build_number = build_number,
        build_time   = build_time,
        status       = status,
    )

    build_sheet(wb,
        sheet_name   = "Board2 - SAMA5D2",
        board_id     = "board2",
        board_label  = "SAMA5D2",
        board_port   = "/dev/ttyUSB1 @ 115200",
        build_number = build_number,
        build_time   = build_time,
        status       = status,
    )

    os.makedirs("test-reports", exist_ok=True)
    out = f"test-reports/board_test_report_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out)
    print(f"✅ Report saved: {out}")


if __name__ == "__main__":
    main()
